import json
import sys 
import os
from openpyxl import load_workbook

script_path = sys.argv[0]
script_dir = os.path.dirname(os.path.abspath(script_path))
medallium_dir = os.path.dirname(script_dir)

name = "yw3"

wb = load_workbook(f"{script_dir}\\{name}.xlsx")
sheet = wb.active

if not "leg" in name:
    dict_ = {row[1]:(str(row[0]),row[2],row[3],row[4]) for row in sheet.iter_rows(values_only=True) if row[0] != None}
else:
    dict_ = {row[0]:(row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8]) for row in sheet.iter_rows(values_only=True) if row[0] != None}

for row in sheet.iter_rows(values_only=True):
    if row[0] != None:
        print(row)

string = json.dumps(dict_)

with open(f"{medallium_dir}\\{name}.json", "w", encoding="utf-8") as f:
    f.write(string)